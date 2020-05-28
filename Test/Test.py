import urllib
from bs4 import BeautifulSoup
import requests
import datetime
from datetime import datetime as dt
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import urllib3 as url
import certifi as cert

Spreadsheet = 'DailyPrice.xlsx'

def get_stock_price(name):
    http = url.PoolManager(cert_reqs='CERT_REQUIRED', ca_certs=cert.where())
    html_doc = http.request('GET', 'https://finance.yahoo.com/quote/' + name)
    soup = BeautifulSoup(html_doc.data, 'html.parser')
    return soup.find("span", class_="Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)").get_text()

#<span class="Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)" data-reactid="50">283.80</span>


if __name__ == '__main__':
    # Fetching currencies
    tsla_price = get_stock_price(CCC)
    aapl_price = get_stock_price(aapl)


    # Writing the scraped data into a spreadsheet
    scraped_data = [dt.strftime(datetime.datetime.now().date(), '%Y/%m/%d'), tsla_price, aapl_price]

    # Spreadsheet Exists, append the scraped data to it
    try:
        wb = load_workbook(Spreadsheet)
        ws = wb.active

        ws.insert_rows(2)
        for i in range(len(scraped_data)):
            ws.cell(row=2, column=i+1).value = scraped_data[i]

        wb.save(Spreadsheet)
        print('Data added to ' + Spreadsheet)

    # Spreadsheet DNE
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(['Datetime', 'Tesla', 'Apple'])
        ws.append(scraped_data)
        wb.save(Spreadsheet)
        print('Data written in ' + Spreadsheet)

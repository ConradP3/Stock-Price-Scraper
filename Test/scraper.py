from urllib.request import urlopen
from bs4 import BeautifulSoup
import urllib3 as url
import certifi as cert
import xlsxwriter
from datetime import datetime as dt
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook


def get_stock_price(name):
    http = url.PoolManager(cert_reqs='CERT_REQUIRED', ca_certs=cert.where())
    html_doc = http.request('GET', 'https://finance.yahoo.com/quote/' + name + '?p=' + name)
    soup = BeautifulSoup(html_doc.data, 'html.parser')
    price = soup.find("span", class_="Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)").get_text()
    return soup.find("span", class_="Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)").get_text()



if __name__ == '__main__':

    stocklist = ['AAPL','AMZN', 'TSLA','SPY', 'VOO', 'VXUS', 'MSFT', 'BAC', 'FB', 'CSCO', 'KO', 'T', 'DAL', 'UAL', 'CCL','F']
    Spreadsheet = 'StockPrice.xlsx'

    #spreadsheet
    tsla_price = get_stock_price('TSLA')
    aapl_price = get_stock_price('AAPL')
    amzn_price = get_stock_price('AMZN')
    spy_price = get_stock_price('SPY')
    voo_price = get_stock_price('VOO')
    scraped_data = [dt.strftime(datetime.datetime.now().date(), '%Y/%m/%d'), tsla_price, aapl_price, amzn_price, spy_price, voo_price]

    print('Symbol | Price')

    for stock in stocklist:
        #Command Line
        print('__________________________________')
        print(str(stock))
        get_stock_price(stock)
        print('     '+ get_stock_price(stock))





    try:
        wb = load_workbook(Spreadsheet)
        ws = wb.active
        ws.insert_rows(2)
        for i in range(len(scraped_data)):
            ws.cell(row=2, column=i+1).value = scraped_data[i]

        wb.save(Spreadsheet)
        print('Data added to ' + Spreadsheet)

    # Spreadsheet DNE
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(['Datetime', 'Tesla', 'Apple', 'Amazon', 'Spy', 'Voo'])
        ws.append(scraped_data)
        wb.save(Spreadsheet)
        print('Data written in ' + Spreadsheet)
